import pandas as pd
import tkinter as tk
from tkinter import filedialog
from math import exp
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Nome colunas
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def add_borders(ws):
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def process_excel(file_path):
    # Lê o arquivo Excel
    df = pd.read_excel(file_path)

    # Coluna A 
    df['Circunferência'] = df.iloc[:, 0] + 0

    # Coluna B
    df['Diâmetro'] = (df['Circunferência'] / 3.1416).round().astype(int)

    # Coluna C
    df['Partes perenes acima e abaixo do solo mais folhas'] = (exp(-6.0032) * df['Diâmetro']**3.7578).round(1)

    # Coluna D
    df['Biomassa C (kg)'] = (df['Partes perenes acima e abaixo do solo mais folhas'] / 2).round(1)

    # Coluna F
    df['CO2eq'] = (0.004198 * df['Diâmetro']**3.779016).round(2)

    # Linhas de média
    mean_row = {
        'Circunferência': 'Média',
        'Diâmetro': df['Diâmetro'].mean().round(),
        'Partes perenes acima e abaixo do solo mais folhas': df['Partes perenes acima e abaixo do solo mais folhas'].mean().round(1),
        'Biomassa C (kg)': df['Biomassa C (kg)'].mean().round(1),
        'CO2eq': df['CO2eq'].mean().round(2)
    }

    mean_df = pd.DataFrame([mean_row])
    df = pd.concat([df, mean_df], ignore_index=True)

    # Seleciona apenas as colunas desejadas para o arquivo de saída
    df_output = df[['Circunferência', 'Diâmetro', 'Partes perenes acima e abaixo do solo mais folhas', 'Biomassa C (kg)', 'CO2eq']]

    # Salva o arquivo Excel 
    output_file = file_path.replace('.xlsx', '_Calculado.xlsx')
    df_output.to_excel(output_file, index=False)

    # Largura das células e adiciona bordas
    wb = load_workbook(output_file)
    ws = wb.active
    adjust_column_width(ws)
    add_borders(ws)

    # Cor do cabeçalho (linha 1) e linha de média (última linha)
    medium_green_fill = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    header_row = ws[1]
    mean_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = medium_green_fill
    for cell in mean_row:
        cell.fill = medium_green_fill

    wb.save(output_file)

    # Gera gráfico
    num_plants = len(df) - 1  # Exclui linha de média
    diameter_classes = ['19 a 22', '22 a 25', '25 a 28', '28 a 31', '31 a 34', '34 a 37', '37 a 40']
    diameter_counts = [((df['Diâmetro'] >= 19) & (df['Diâmetro'] <= 22)).sum(),
                       ((df['Diâmetro'] > 22) & (df['Diâmetro'] <= 25)).sum(),
                       ((df['Diâmetro'] > 25) & (df['Diâmetro'] <= 28)).sum(),
                       ((df['Diâmetro'] > 28) & (df['Diâmetro'] <= 31)).sum(),
                       ((df['Diâmetro'] > 31) & (df['Diâmetro'] <= 34)).sum(),
                       ((df['Diâmetro'] > 34) & (df['Diâmetro'] <= 37)).sum(),
                       ((df['Diâmetro'] > 37) & (df['Diâmetro'] <= 40)).sum()]

    plt.bar(diameter_classes, diameter_counts)
    plt.xlabel('Classes de Diâmetro')
    plt.ylabel('Nº de Plantas')
    plt.title('Distribuição do Diâmetro das Plantas')
    plt.xticks(rotation=45, ha='right')
    plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.3)  # Ajusta margens
    plt.savefig(file_path.replace('.xlsx', '_Graph.png'))  # Salva o gráfico como uma imagem
    plt.close()  # Fecha a figura após salvar

    # Mensagem de sucesso
    success_label.config(text=f"Arquivo processado salvo como {output_file}")

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        process_excel(file_path)

# Interface gráfica
root = tk.Tk()
root.title("CaCrEst")

frame = tk.Frame(root, padx=20, pady=20)
frame.pack(padx=20, pady=20)

select_button = tk.Button(frame, text="Selecionar Arquivo Excel", command=select_file)
select_button.pack()

success_label = tk.Label(frame, text="")
success_label.pack()

root.geometry("580x200")  # Tamanho do painel
root.mainloop()
